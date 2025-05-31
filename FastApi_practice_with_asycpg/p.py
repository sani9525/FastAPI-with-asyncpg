from openpyxl import Workbook
import os

def save_to_excel(user_data, file_name="users.xlsx"):
    file_exists = os.path.exists(file_name)

    if not file_exists:
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Email", "Password"])  # header
    else:
        from openpyxl import load_workbook
        wb = load_workbook(file_name)
        ws = wb.active

    ws.append([user_data["name"], user_data["email"], user_data["password"]])
    wb.save(file_name)




user = {"name": "Sani", "email": "sani@example.com", "password": "securepass"}
save_to_excel(user)
